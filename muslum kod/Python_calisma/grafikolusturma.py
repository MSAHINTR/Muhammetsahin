from matplotlib import pyplot as plt
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import xlrd
import pandas
import os
#Girdi araçları
print("lütfen dosyanızın scriptin olduğu klasorde oldugundan emin olunuz")
dosyayolu=input("lütfen dosya adını giriniz: ")
#Dosya adını girmeniz yeterli uzantı xlsx olarak otonmatik atnacaktır 

dosyayolu=dosyayolu+".xlsx"

#İstenildiği koşulda herhangi bir x degeri alınıp bu x değerine göre grafikte yer gosterici açılabilir. 

#secilmisxdegeri=float (input("istenilen x mesafesindeki değeri giriniz: "))

#stun yerini temel verilerin oldugu yerden uzaklaştırıyoruzki birisi asıl verilerin üstüne yanlışlıkla birşeyler yazdırmasın

sutunnyeri=3             #int(input("yazıalacak stunu belirtin "))

#eger stun yerini belirtmek isterseniz lütfen bu kısmı ve 13. satırdaki input kısmını devreye sokun 
"""
if sutunnyeri==1:
    sutunnyeri+=2
if sutunnyeri==1:
    sutunnyeri+=1
if sutunnyeri<0:
    sutunnyeri=-1*sutunnyeri
"""
wb = load_workbook(dosyayolu)

#Sheetname kısmını programda veya excel de girmek gerekir. Excel dosyasında sayfa adı değişmediyse burayı degiştireye gerek yok. 

sheetnamee = ('Sheet1')              #input("lütfen aktif çalışma sayfasının adını giriniz")
# Sayfa adını cagırma 

sheet = wb[sheetnamee]

# Dosyadaki satir ve stun sayısını cekme 

inputWorkbook=xlrd.open_workbook(dosyayolu)
inputWorksheet = inputWorkbook.sheet_by_index(0)

satirsayisi=int(inputWorksheet.nrows)

#ilgili calışma sayfasını belirleme ve bu sayfaya veri girmek için sayfayı getirme 
sayfa=wb.get_sheet_by_name(sheetnamee)

# okunan değerleri işleme 

df=pd.read_excel(dosyayolu,sheetnamee)

stunlar=np.array(df[0:],dtype=np.float)
x1=stunlar[:,0]/10000
y1=stunlar[:,1]

#Paşam bak burası acayip çokomelli senin x ekseninde tekrar donuş yaptığın noktayı belirliyor. Bu nokta kayarsa hayatımız kayar :P

a=0
k=0
l=1

while a<satirsayisi:
   if 1<x1[k]>x1[l]:
       k+=2
       a=k
       break
   else:
       k+=1
       l+=1
#satır sayısı tanımlamaları (birkaç gereksiz sabit olabilir aldırma)
sutunn=sutunnyeri
a=k
secilmisydegeri=0
b=a+2
#satır stun yazmayagerek yok dedin yazdırmadık bak tamamen isteğe özel 
#sayfa.cell(row=1,column=sutunnyeri, value=stunadi)
#Paşam burası bak çokomelli fazla kurcalamaya gelmez hemen bozuluyor :P

while satirsayisi-1<satirsayisi:
    if a<2:
        break
    elif b>(satirsayisi):
        break
    elif float(x1[a])<float(x1[b]):
        b+=1
    elif float(x1[a])>float(x1[b]):
        
        p=float(2*(y1[b]-y1[a])/18)
        #x degelrerini buradan excel'e atıyoruz, Burada p satırdaki degeri gösterir
        sayfa.cell(row=a+2,column=sutunn, value=p)
        a-=1
#Excel dosyasına geri beslemedir. dosyayı kaydedip çıkış yapar sonra tekrar okumak için açıp verilerin kaybolmasını engeller. 
wb.save(dosyayolu)
wb.close
df=pd.read_excel(dosyayolu,sheetnamee)

#Normalde gorunmeyen 3. stun dosyayı kapatıp actığımızdan gorunur oldu. ve bu değeride hemen çekiyoruz bu şekilde y ekseni hazır.

hesaplanmisveri=stunlar[:,2]

#   Bu kısım y kordinatını bulmaya yarar 

a=k
b=k+1
x=1
#İt's working clearly. But it's not wanting by muslum guven:P
"""while x<(x+1):
    if float(x1[a])>float(secilmisxdegeri):
        a-=1
    elif float(x1[b])>float(secilmisxdegeri):
        b+=1
    else:
        p=float(2*(y1[b]-y1[a])/18)
        secilmisydegeri=p
        print (a)
        print (b)
        break
"""
#Grafik oluşturma kısmı

plt.rcParams['figure.figsize'] = [6, 4]
figure=plt.figure()
plt.figure (2,dpi=600)
plt.subplot(211)

plt.plot(x1,y1,label="Sonuclar")
plt.ylabel(df.columns[1])
plt.xticks(np.arange(-4,9,1))
plt.yticks(np.arange(-6,6,1.4))
plt.grid()
plt.title("calculation",loc="center")

plt.subplot(212)
plt.plot(x1,hesaplanmisveri,color='green')
plt.xlabel(df.columns[0])
plt.ylabel(df.columns[2])
plt.xticks(np.arange(-4,9,1))
plt.yticks(np.arange(-0.2,1.3,0.2))
plt.grid()
#plt.annotate("zirve", (40,0.5), (40,0,5), arrowprops=dict(arrowstyle="->"))
#plt.arrow(x=secilmisxdegeri, y=secilmisydegeri+0.2, dx=0, dy=-secilmisydegeri-0.3, width=0.001) 
#plt.annotate(secilmisydegeri + "Am^-2", xy = (secilmisxdegeri+0.5, secilmisydegeri-0.2))
#kaydetme 
kayidedilecekdosyadi=input("Kaydetmek istediğiniz dosya adını giriniz")
plt.savefig(kayidedilecekdosyadi)
os.startfile(kayidedilecekdosyadi+'.png')
